from django.contrib.auth.hashers import make_password
from ..models import VoterUserMaster, UploadedLoginExcel
import re
from django.db import transaction
from openpyxl import load_workbook
import base64
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.permissions import AllowAny,IsAuthenticated
from application.models import VoterUserMaster
import io
from rest_framework.decorators import api_view, permission_classes, parser_classes
from rest_framework.response import Response
from logger import logger
from rest_framework import status
from .view_utils import validate_password

def is_valid_mobile(mobile):
    # Allows only exactly 10 digits
    pattern = r'^[6-9]\d{9}$'
    return bool(re.match(pattern, mobile))

# ---------- DROP-IN HELPER (REQUIRED) ----------
def normalize_mobile(mobile):
    if mobile is None:
        return None

    if isinstance(mobile, float):
        return str(int(mobile))

    if isinstance(mobile, int):
        return str(mobile)

    mobile = str(mobile).strip()

    if mobile.endswith(".0"):
        mobile = mobile[:-2]

    return mobile



@api_view(["POST"])
@permission_classes([AllowAny])
def registration(request):
    data = request.data
    logger.info("registration_api: Registration request received")

    first_name = data.get("first_name")
    last_name = data.get("last_name")
    mobile_no = data.get("mobile_no")
    password = data.get("password")
    confirm_password = data.get("confirm_password")
    role_str = data.get("role")  # optional

    # ---------- BASIC VALIDATIONS ----------
    if not first_name:
        return Response({"status": False, "message": "First Name is required"}, status=400)

    if not last_name:
        return Response({"status": False, "message": "Last Name is required"}, status=400)

    if not mobile_no or not is_valid_mobile(mobile_no):
        return Response({"status": False, "message": "Invalid mobile number"}, status=400)

    if VoterUserMaster.objects.filter(mobile_no=mobile_no).exists():
        return Response({"status": False, "message": "Mobile already registered"}, status=400)

    # ---------- PASSWORD VALIDATION ----------
    try:
        validate_password(
            new_password=password,
            confirm_password=confirm_password,
            phone=mobile_no,
            user=None        # registration = no existing user
        )
    except ValueError as e:
        return Response(
            {"status": False, "message": str(e)},
            status=status.HTTP_400_BAD_REQUEST
        )

    # ---------- ROLE LOGIC ----------
    ROLE_MAP = {
        "superadmin": 1,
        "admin": 2,
        "volunteer": 3,
    }

    role_id = ROLE_MAP.get(role_str.lower(), 3) if role_str else 3

    # ---------- CREATE USER ----------
    user = VoterUserMaster.objects.create(
        first_name=first_name,
        last_name=last_name,
        mobile_no=mobile_no,
        password=make_password(password),
        role_id=role_id
    )

    logger.info(
        f"registration_api: User {user.user_id} registered successfully with role ID {role_id}"
    )

    # ---------- RESPONSE ----------
    return Response(
        {
            "status": True,
            "message": "Registration successful",
            "data": {
                "user_id": user.user_id,
                "first_name": user.first_name,
                "last_name": user.last_name,
                "mobile_no": user.mobile_no,
                "role_id": role_id
            }
        },
        status=status.HTTP_201_CREATED
    )
  

@api_view(["POST"])
@permission_classes([IsAuthenticated])
@parser_classes([MultiPartParser, FormParser])
def upload_login_credentials_excel(request):

    try:
        logger.info("upload_login_credentials_excel_api: Upload login credentials Excel request received")
        excel_file = request.FILES.get("file")

        if not excel_file:
            return Response(
                {"status": False, "message": "Excel file is required"},
                status=400
            )

        file_name = excel_file.name
        file_bytes = excel_file.read()
        file_base64 = base64.b64encode(file_bytes).decode("utf-8")

        # -------- LOAD EXCEL --------
        wb = load_workbook(io.BytesIO(file_bytes))
        sheet = wb.active

        raw_headers = [
            str(cell.value).strip().lower()
            for cell in sheet[1]
            if cell.value
        ]

        required_headers = {
            "first name",
            "last name",
            "mobile number",
            "password"
        }

        if not required_headers.issubset(set(raw_headers)):
            return Response(
                {
                    "status": False,
                    "message": "Invalid Excel format",
                    "required_columns": list(required_headers)
                },
                status=400
            )

        created = 0
        skipped = 0
        errors = []

        # -------- SAVE EXCEL META --------
        uploaded_excel = UploadedLoginExcel.objects.create(
            file_name=file_name,
            file_base64=file_base64
        )

        header_index = {h: i for i, h in enumerate(raw_headers)}

        # -------- PROCESS ROWS --------
        with transaction.atomic():
            for row_no, row in enumerate(
                sheet.iter_rows(min_row=2, values_only=True),
                start=2
            ):

                first_name = row[header_index["first name"]]
                last_name = row[header_index["last name"]]
                mobile_no = row[header_index["mobile number"]]
                password = row[header_index["password"]]

                # ---------- DROP-IN FIX ----------
                mobile_no = normalize_mobile(mobile_no)

                # ---------- HARD STOP ----------
                if not mobile_no or not mobile_no.isdigit() or len(mobile_no) != 10:
                    skipped += 1
                    errors.append(
                        f"Row {row_no}: invalid mobile number ({mobile_no})"
                    )
                    continue

                if not first_name or not password:
                    skipped += 1
                    errors.append(f"Row {row_no}: missing required fields")
                    continue

                if VoterUserMaster.objects.filter(mobile_no=mobile_no).exists():
                    skipped += 1
                    errors.append(
                        f"Row {row_no}: mobile already exists ({mobile_no})"
                    )
                    continue

                # ---------- INSERT ----------
                VoterUserMaster.objects.create(
                    first_name=str(first_name).strip(),
                    last_name=str(last_name).strip() if last_name else "",
                    mobile_no=mobile_no,
                    password=make_password(str(password)),
                    role_id=3
                )

                created += 1

        uploaded_excel.created_count = created
        uploaded_excel.skipped_count = skipped
        uploaded_excel.save()
        logger.info(f"upload_login_credentials_excel_api: Excel imported successfully with {created} created users and {skipped} skipped rows")
        return Response({
            "status": True,
            "message": "Excel imported successfully",
            "excel_id": uploaded_excel.id,
            "created_users": created,
            "skipped_rows": skipped,
            "errors": errors
        })

    except Exception as e:
        return Response(
            {"status": False, "error": str(e)},
            status=500
        )


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def list_uploaded_login_excels(request):
    excels = UploadedLoginExcel.objects.order_by("-uploaded_at")
    logger.info("list_uploaded_login_excels_api: List uploaded login excels request received")
    data = []
    for e in excels:
        data.append({
            "excel_id": e.id,
            "file_name": e.file_name,
            "uploaded_at": e.uploaded_at.strftime("%d-%m-%Y %I:%M %p"),
            "created_users": e.created_count,
            "skipped_rows": e.skipped_count
        })
    logger.info(f"list_uploaded_login_excels_api: Retrieved {len(data)} uploaded login excels")
    return Response({
        "status": True,
        "count": len(data),
        "data": data
    })
    
@api_view(["DELETE"])
@permission_classes([IsAuthenticated])
def delete_uploaded_login_excel(request, excel_id):

    try:
        logger.info(f"delete_uploaded_login_excel_api: Delete uploaded login excel request received for excel_id={excel_id}")
        excel = UploadedLoginExcel.objects.get(id=excel_id)
    except UploadedLoginExcel.DoesNotExist:
        return Response(
            {
                "status": False,
                "message": "Excel record not found"
            },
            status=404
        )

    excel.delete()
    logger.info(f"delete_uploaded_login_excel_api: Excel record with excel_id={excel_id} deleted successfully")
    return Response(
        {
            "status": True,
            "message": "Excel record deleted successfully",
            "excel_id": excel_id
        },
        status=200
    )

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def download_login_excel(request, excel_id):
    excel = get_object_or_404(UploadedLoginExcel, id=excel_id)

    base64_data = excel.file_base64

    # remove data URI prefix if present
    if "," in base64_data:
        base64_data = base64_data.split(",", 1)[1]

    # remove whitespace / newlines
    base64_data = "".join(base64_data.split())

    #  DECODE EXACTLY ONCE
    file_bytes = base64.b64decode(base64_data)

    #  sanity check (must be ZIP)
    if not file_bytes.startswith(b"PK"):
        return Response(
            {"status": False, "message": "Not a valid Excel file"},
            status=400
        )

    response = HttpResponse(
        file_bytes,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    filename = excel.file_name
    if not filename.lower().endswith(".xlsx"):
        filename += ".xlsx"

    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    response["Content-Length"] = len(file_bytes)

    return response